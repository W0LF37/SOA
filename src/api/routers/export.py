from __future__ import annotations

import io
import json
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.core.runtime_paths import prepare_writable_file_path

PROJECT_ROOT = Path(__file__).resolve().parents[3]
TASKS_PATH = prepare_writable_file_path(PROJECT_ROOT, "data/processed/tasks.json")
FINAL_TASKS_PATH = prepare_writable_file_path(PROJECT_ROOT, "data/processed/tasks_final.json")

router = APIRouter()


@router.get("/tasks")
def export_tasks_excel() -> StreamingResponse:
    tasks_path = FINAL_TASKS_PATH if FINAL_TASKS_PATH.exists() else TASKS_PATH
    if not tasks_path.exists():
        raise HTTPException(status_code=404, detail="No plan generated yet")

    payload = json.loads(tasks_path.read_text(encoding="utf-8"))
    tasks = payload.get("tasks", [])
    if not isinstance(tasks, list):
        tasks = []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = [
        "ID",
        "Title",
        "Type",
        "Type Reason",
        "Complexity",
        "Complexity Reason",
        "Estimated Hours",
        "Dependencies",
        "Confidence",
        "Owner Role",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")

    for row, task in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=task.get("id", ""))
        ws.cell(row=row, column=2, value=task.get("title", ""))
        ws.cell(row=row, column=3, value=task.get("req_type", ""))
        ws.cell(row=row, column=4, value=task.get("type_reason", ""))
        ws.cell(row=row, column=5, value=task.get("complexity", ""))
        ws.cell(row=row, column=6, value=task.get("complexity_reason", ""))
        ws.cell(row=row, column=7, value=task.get("estimated_hours", ""))
        ws.cell(
            row=row,
            column=8,
            value=", ".join(task.get("dependencies", [])),
        )
        ws.cell(row=row, column=9, value=task.get("confidence", ""))
        ws.cell(row=row, column=10, value=task.get("suggested_owner_role", ""))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=critiplan_tasks.xlsx"
        },
    )
